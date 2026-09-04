"""Entry lists arriving as a spreadsheet.

The load-bearing test here is the round trip: download the template, fill it
in, upload it back. That is the machine-checkable form of "the template and the
parser cannot drift apart", which is the whole reason both read one column
definition.

The rest pin down that the preview writes nothing, that the commit re-validates
what the browser sends rather than trusting it, and one error case per code.
"""

import asyncio
from datetime import date
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.core import database as database_module
from app.main import app
from app.models.base import Base
from app.modules.tournaments.services.participant_import import IMPORT_COLUMNS, SHEET_ENTRIES

EVENT_YEAR = 2026
START_DATE = date(EVENT_YEAR, 5, 16).isoformat()
XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def setup_app_for_tests():
    engine = create_async_engine("sqlite+aiosqlite:///:memory:", future=True)
    database_module.engine = engine
    database_module.AsyncSessionLocal = async_sessionmaker(
        bind=engine, class_=AsyncSession, expire_on_commit=False
    )

    async def setup_db() -> None:
        async with engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)

    asyncio.run(setup_db())

    def override_get_db():
        async def _override():
            async with database_module.AsyncSessionLocal() as session:
                yield session

        return _override

    app.dependency_overrides[database_module.get_db] = override_get_db()
    return TestClient(app)


def register(client, email: str) -> tuple[str, dict[str, str]]:
    response = client.post(
        "/api/v1/auth/register",
        json={"email": email, "password": "StrongPassword123!", "first_name": "Иван", "last_name": "Организатор"},
    )
    assert response.status_code == 201, response.text
    headers = {"Authorization": f"Bearer {response.json()['access_token']}"}
    me = client.get("/api/v1/users/me", headers=headers)
    return me.json()["id"], headers


def bootstrap(client):
    """A tournament with three disciplines, one of them age-bounded."""
    organizer_id, headers = register(client, "organizer@example.com")
    ruleset = client.post("/api/v1/rulesets", json={"title": "Base", "version": "1.0", "status": "ACTIVE"})
    tournament = client.post(
        "/api/v1/tournaments",
        json={
            "title": "Мстинская традиция 2026",
            "status": "REGISTRATION",
            "start_date": START_DATE,
            "organizer_id": organizer_id,
            "ruleset_id": ruleset.json()["id"],
        },
    )
    assert tournament.status_code == 201, tournament.text
    tournament_id = tournament.json()["id"]

    for name, extra in (
        ("Абсолютная взрослая", {}),
        ("Ветераны", {"min_age": 45}),
        ("Абсолютная детская", {"max_age": 14}),
    ):
        created = client.post(
            f"/api/v1/tournaments/{tournament_id}/competitions",
            json={
                "tournament_id": tournament_id,
                "name": name,
                "type": "INDIVIDUAL",
                "format": "SINGLE_ELIMINATION",
                "status": "REGISTRATION",
                **extra,
            },
        )
        assert created.status_code == 201, created.text
    return tournament_id, headers


def sheet_of(rows: list[dict]) -> bytes:
    """Build an .xlsx the way an organizer would, from the real headers."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_ENTRIES
    sheet.append([column.header_ru for column in IMPORT_COLUMNS])
    for row in rows:
        sheet.append([row.get(column.key, "") for column in IMPORT_COLUMNS])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def preview(client, tournament_id: str, payload: bytes, headers: dict[str, str]):
    return client.post(
        f"/api/v1/tournaments/{tournament_id}/participants/import/preview",
        files={"file": ("entries.xlsx", payload, XLSX)},
        headers=headers,
    )


def participants(client, tournament_id: str) -> list[dict]:
    competitions = client.get(f"/api/v1/tournaments/{tournament_id}/competitions").json()
    everyone: list[dict] = []
    for competition in competitions:
        everyone += client.get(f"/api/v1/competitions/{competition['id']}/participants").json()
    return everyone


# --------------------------------------------------------------- round trip


def test_the_template_can_be_filled_in_and_uploaded_back():
    """The anti-drift guarantee, checked rather than asserted in a comment."""
    client = setup_app_for_tests()
    tournament_id, headers = bootstrap(client)

    downloaded = client.get(
        f"/api/v1/tournaments/{tournament_id}/participants/template.xlsx", headers=headers
    )
    assert downloaded.status_code == 200, downloaded.text
    assert downloaded.headers["content-type"].startswith(XLSX)

    workbook = load_workbook(BytesIO(downloaded.content))
    sheet = workbook[SHEET_ENTRIES]
    headers_row = [cell.value for cell in sheet[1]]
    assert headers_row == [column.header_ru for column in IMPORT_COLUMNS]

    # The second sheet tells the organizer which category names are accepted.
    reference = workbook["Дисциплины"]
    names = [row[0] for row in reference.iter_rows(min_row=2, values_only=True)]
    assert "Ветераны" in names

    # Fill the downloaded file in and send it straight back.
    sheet.append(["Замятин Пётр", "Кистень", "Новгород", "Буза", "Ветераны", EVENT_YEAR - 50, ""])
    filled = BytesIO()
    workbook.save(filled)

    report = preview(client, tournament_id, filled.getvalue(), headers)
    assert report.status_code == 200, report.text
    body = report.json()
    assert body["total_rows"] == 1, body["rows"]
    assert body["valid_rows"] == 1
    assert body["rows"][0]["competition_name"] == "Ветераны"


# ------------------------------------------------------- preview writes nothing


def test_the_preview_persists_nothing():
    client = setup_app_for_tests()
    tournament_id, headers = bootstrap(client)

    payload = sheet_of(
        [{"full_name": "Иван Иванов", "category": "Абсолютная взрослая", "city": "Псков"}]
    )
    assert preview(client, tournament_id, payload, headers).json()["valid_rows"] == 1
    assert participants(client, tournament_id) == []


# ---------------------------------------------------------------- the commit


def test_a_reviewed_list_is_entered_and_journalled():
    client = setup_app_for_tests()
    tournament_id, headers = bootstrap(client)

    payload = sheet_of(
        [
            {
                "full_name": "Замятин Пётр",
                "fight_name": "Кистень",
                "city": "Новгород",
                "club": "Буза",
                "category": "Абсолютная взрослая",
                "seed": 1,
            },
            {"full_name": "Сергеев Сергей", "city": "Псков", "category": "Абсолютная взрослая"},
        ]
    )
    report = preview(client, tournament_id, payload, headers).json()
    assert report["valid_rows"] == 2

    committed = client.post(
        f"/api/v1/tournaments/{tournament_id}/participants/import/commit",
        json={"rows": report["rows"]},
        headers=headers,
    )
    assert committed.status_code == 200, committed.text
    assert committed.json()["created"] == 2
    assert committed.json()["per_competition"] == {"Абсолютная взрослая": 2}

    entered = participants(client, tournament_id)
    by_name = {row["display_name"]: row for row in entered}
    # Драковое имя wins where there is one; ФИО where there is not.
    assert set(by_name) == {"Кистень", "Сергеев Сергей"}
    assert by_name["Кистень"]["club_name"] == "Буза"
    assert by_name["Кистень"]["city"] == "Новгород"
    assert by_name["Кистень"]["seed"] == 1

    competition_id = by_name["Кистень"]["competition_id"]
    journal = client.get(f"/api/v1/competitions/{competition_id}/events").json()
    assert [e for e in journal if e["event_type"] == "PARTICIPANTS_IMPORTED"]


def test_the_commit_revalidates_what_the_browser_sends():
    """A row edited to something invalid after the preview is still refused."""
    client = setup_app_for_tests()
    tournament_id, headers = bootstrap(client)

    payload = sheet_of([{"full_name": "Иван Иванов", "category": "Абсолютная взрослая"}])
    report = preview(client, tournament_id, payload, headers).json()
    assert report["valid_rows"] == 1

    tampered = report["rows"][0] | {"category": "Ветераны", "birth_year": EVENT_YEAR - 20}
    refused = client.post(
        f"/api/v1/tournaments/{tournament_id}/participants/import/commit",
        json={"rows": [tampered]},
        headers=headers,
    )
    assert refused.status_code == 400, refused.text
    assert participants(client, tournament_id) == []


def test_a_batch_with_one_bad_row_is_refused_whole():
    """Half an entry list is worse than none — the organizer cannot tell which half."""
    client = setup_app_for_tests()
    tournament_id, headers = bootstrap(client)

    payload = sheet_of(
        [
            {"full_name": "Хороший", "category": "Абсолютная взрослая"},
            {"full_name": "", "category": "Абсолютная взрослая"},
        ]
    )
    report = preview(client, tournament_id, payload, headers).json()
    refused = client.post(
        f"/api/v1/tournaments/{tournament_id}/participants/import/commit",
        json={"rows": report["rows"]},
        headers=headers,
    )
    assert refused.status_code == 400, refused.text
    assert participants(client, tournament_id) == []


# ------------------------------------------------------------ the error codes


def codes(report: dict, row_index: int = 0) -> set[str]:
    return {error["code"] for error in report["rows"][row_index]["errors"]}


def test_a_missing_name_is_reported():
    client = setup_app_for_tests()
    tournament_id, headers = bootstrap(client)
    report = preview(
        client, tournament_id, sheet_of([{"full_name": "", "category": "Ветераны"}]), headers
    ).json()
    assert "MISSING_NAME" in codes(report)


def test_an_unknown_category_is_reported_once_for_the_file():
    client = setup_app_for_tests()
    tournament_id, headers = bootstrap(client)
    report = preview(
        client,
        tournament_id,
        sheet_of(
            [
                {"full_name": "Первый", "category": "Женская абсолютка"},
                {"full_name": "Второй", "category": "Женская абсолютка"},
            ]
        ),
        headers,
    ).json()
    assert "UNKNOWN_CATEGORY" in codes(report)
    assert report["unknown_categories"] == ["Женская абсолютка"]


def test_the_age_bound_is_checked_on_import_too():
    client = setup_app_for_tests()
    tournament_id, headers = bootstrap(client)
    report = preview(
        client,
        tournament_id,
        sheet_of(
            [
                {"full_name": "Молодой", "category": "Ветераны", "birth_year": EVENT_YEAR - 30},
                {"full_name": "Без года", "category": "Ветераны"},
                {"full_name": "Взрослый", "category": "Абсолютная детская", "birth_year": EVENT_YEAR - 30},
            ]
        ),
        headers,
    ).json()
    assert "AGE_BELOW_MINIMUM" in codes(report, 0)
    assert "MISSING_BIRTH_YEAR" in codes(report, 1)
    assert "AGE_ABOVE_MAXIMUM" in codes(report, 2)


def test_a_bad_year_or_seed_is_reported():
    client = setup_app_for_tests()
    tournament_id, headers = bootstrap(client)
    report = preview(
        client,
        tournament_id,
        sheet_of(
            [
                {
                    "full_name": "Кривой",
                    "category": "Абсолютная взрослая",
                    "birth_year": "позапрошлый",
                    "seed": "первый",
                }
            ]
        ),
        headers,
    ).json()
    assert {"INVALID_BIRTH_YEAR", "INVALID_SEED"} <= codes(report)


def test_a_duplicate_inside_the_file_is_reported():
    client = setup_app_for_tests()
    tournament_id, headers = bootstrap(client)
    report = preview(
        client,
        tournament_id,
        sheet_of(
            [
                {"full_name": "Иван Иванов", "category": "Абсолютная взрослая"},
                {"full_name": "Иван  Иванов", "category": "Абсолютная взрослая"},
            ]
        ),
        headers,
    ).json()
    # The first occurrence is fine; the second is the duplicate.
    assert codes(report, 0) == set()
    assert "DUPLICATE_IN_FILE" in codes(report, 1)


def test_someone_already_entered_is_reported():
    client = setup_app_for_tests()
    tournament_id, headers = bootstrap(client)
    competitions = client.get(f"/api/v1/tournaments/{tournament_id}/competitions").json()
    absolute = next(c for c in competitions if c["name"] == "Абсолютная взрослая")
    client.post(
        f"/api/v1/competitions/{absolute['id']}/participants",
        json={"competition_id": absolute["id"], "display_name": "Иван Иванов"},
    ).raise_for_status()

    report = preview(
        client,
        tournament_id,
        sheet_of([{"full_name": "Иван Иванов", "category": "Абсолютная взрослая"}]),
        headers,
    ).json()
    assert "DUPLICATE_IN_COMPETITION" in codes(report)


def test_the_same_person_may_be_entered_in_two_disciplines():
    """Not a duplicate: «Ветераны» and the open absolute are different fields."""
    client = setup_app_for_tests()
    tournament_id, headers = bootstrap(client)
    report = preview(
        client,
        tournament_id,
        sheet_of(
            [
                {"full_name": "Замятин Пётр", "category": "Ветераны", "birth_year": EVENT_YEAR - 50},
                {"full_name": "Замятин Пётр", "category": "Абсолютная взрослая"},
            ]
        ),
        headers,
    ).json()
    assert report["valid_rows"] == 2, report["rows"]


# -------------------------------------------------------- file-level refusals


def test_a_reordered_sheet_still_imports():
    """Columns are matched by header text, not by position."""
    client = setup_app_for_tests()
    tournament_id, headers = bootstrap(client)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_ENTRIES
    sheet.append(["Категория", "Город", "ФИО"])
    sheet.append(["Абсолютная взрослая", "Тверь", "Фёдоров Фёдор"])
    stream = BytesIO()
    workbook.save(stream)

    report = preview(client, tournament_id, stream.getvalue(), headers).json()
    assert report["valid_rows"] == 1
    assert report["rows"][0]["full_name"] == "Фёдоров Фёдор"
    assert report["rows"][0]["city"] == "Тверь"


def test_a_file_that_is_not_a_spreadsheet_is_refused_clearly():
    client = setup_app_for_tests()
    tournament_id, headers = bootstrap(client)
    refused = preview(client, tournament_id, b"not a workbook at all", headers)
    assert refused.status_code == 400, refused.text
    assert "xlsx" in refused.json()["detail"].lower()


# ------------------------------------------------------------------- guards


def test_the_intake_requires_an_authorized_manager():
    client = setup_app_for_tests()
    tournament_id, _ = bootstrap(client)

    anonymous = client.get(f"/api/v1/tournaments/{tournament_id}/participants/template.xlsx")
    assert anonymous.status_code == 401, anonymous.text

    _, stranger = register(client, "stranger@example.com")
    forbidden = client.get(
        f"/api/v1/tournaments/{tournament_id}/participants/template.xlsx", headers=stranger
    )
    assert forbidden.status_code == 403, forbidden.text

    payload = sheet_of([{"full_name": "Чужой", "category": "Абсолютная взрослая"}])
    assert preview(client, tournament_id, payload, stranger).status_code == 403
