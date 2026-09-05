"""Entry lists arriving as a spreadsheet.

The organizer downloads a template, fills it in, uploads it back, reads a
per-row report, fixes what is wrong, and only then commits. Three properties
matter and each is a deliberate choice here:

* **One definition of the columns.** :data:`IMPORT_COLUMNS` is walked by both
  the template writer and the parser, so the file the organizer is handed and
  the file the server expects cannot drift apart. The round-trip test proves it.
* **Nothing is written by the preview.** Validation reads the database — it has
  to, to spot a duplicate or an unknown discipline — but never writes to it.
* **The commit re-validates.** It takes rows, not the file again, because the
  organizer may have corrected a discipline or a birth year in the review table;
  re-parsing would throw those edits away. Which means the submitted rows are
  client-supplied, so they go through the identical validator a second time.

Columns are matched by header text rather than position, so a reordered sheet
still imports.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
from typing import BinaryIO, Iterable, Sequence
from uuid import UUID

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.athletes.models import Athlete
from app.modules.tournaments.domain import eligibility
from app.modules.tournaments.models import (
    Competition,
    CompetitionEvent,
    Participant,
    Tournament,
    TournamentCategory,
)

#: A hand-filled sheet larger than this is a mistake, not an entry list, and
#: parsing it would tie up a worker for no good reason.
MAX_UPLOAD_BYTES = 2 * 1024 * 1024
MAX_DATA_ROWS = 2000

SHEET_ENTRIES = "Участники"
SHEET_DISCIPLINES = "Дисциплины"


@dataclass(frozen=True)
class ImportColumn:
    key: str
    header_ru: str
    required: bool
    note: str
    width: int


#: The one place the file format is described. Order here is the order in the
#: template; the parser matches on ``header_ru`` and only falls back to this
#: order when it recognizes no header at all.
IMPORT_COLUMNS: tuple[ImportColumn, ...] = (
    ImportColumn("full_name", "ФИО", True, "обязательно", 32),
    ImportColumn(
        "fight_name",
        "Драковое имя",
        False,
        "если есть; пусто — в сетке будет ФИО",
        22,
    ),
    ImportColumn("city", "Город", False, "разводит земляков в первом круге", 20),
    ImportColumn("club", "Клуб", False, "разводит одноклубников, важнее города", 24),
    ImportColumn(
        "category",
        "Категория",
        True,
        "название дисциплины из листа «Дисциплины»",
        28,
    ),
    ImportColumn(
        "birth_year",
        "Год рождения",
        False,
        "обязателен только там, где у дисциплины есть возрастное ограничение",
        14,
    ),
    ImportColumn("seed", "Посев", False, "необязательно", 10),
    ImportColumn(
        "reserve",
        "Запасной",
        False,
        "«да» — в сетку не попадёт, ждёт замены",
        12,
    ),
)

COLUMN_BY_KEY = {column.key: column for column in IMPORT_COLUMNS}

#: What counts as "yes" in «Запасной». Anything outside these two sets is a row
#: error rather than a silent "no": a typo in this column decides whether a
#: fighter is in the draw at all.
RESERVE_YES = frozenset({"да", "yes", "y", "д", "1", "+", "true", "истина"})
RESERVE_NO = frozenset({"", "нет", "no", "n", "н", "0", "-", "false", "ложь"})

#: Prefix marking a row the template filled in for illustration. The parser
#: drops these, so a file uploaded without deleting them enters nobody. A
#: marker rather than a row number, because an organizer who inserts a row
#: above would otherwise shift the examples into their own entry list.
EXAMPLE_MARKER = "ПРИМЕР:"

#: Shown filled in so the format is read off a real row rather than guessed at
#: from the notes. Deliberately exercising the awkward parts: a fighter with no
#: драковое имя, and one held in reserve.
EXAMPLE_ROWS: tuple[dict[str, str], ...] = (
    {
        "full_name": f"{EXAMPLE_MARKER} Замятин Пётр Ильич",
        "fight_name": "Кистень",
        "city": "Великий Новгород",
        "club": "Буза",
        "category": "название из листа «Дисциплины»",
        "birth_year": "1998",
        "seed": "1",
        "reserve": "нет",
    },
    {
        "full_name": f"{EXAMPLE_MARKER} Сергеев Сергей Сергеевич",
        "fight_name": "",
        "city": "Псков",
        "club": "Сокол",
        "category": "название из листа «Дисциплины»",
        "birth_year": "2005",
        "seed": "",
        "reserve": "да",
    },
)


def _normalize(value: str | None) -> str:
    """Comparison key for a header, a name or a category — same rule everywhere."""
    if not value:
        return ""
    return " ".join(str(value).split()).casefold()


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


# --------------------------------------------------------------- the template


def build_template_workbook(competitions: Sequence[Competition]) -> BytesIO:
    """The .xlsx the organizer fills in.

    Carries a second, read-only sheet listing the tournament's disciplines with
    their age bounds, so «Категория» can be filled in correctly without
    guessing at a name the parser will then reject.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_ENTRIES

    for index, column in enumerate(IMPORT_COLUMNS, start=1):
        letter = get_column_letter(index)
        header = sheet.cell(row=1, column=index, value=column.header_ru)
        header.font = Font(bold=True)
        note = sheet.cell(row=2, column=index, value=column.note)
        note.font = Font(italic=True, size=9)
        note.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.column_dimensions[letter].width = column.width

    example_font = Font(italic=True, color="FF808080")
    for example in EXAMPLE_ROWS:
        sheet.append([example.get(column.key, "") for column in IMPORT_COLUMNS])
        for cell in sheet[sheet.max_row]:
            cell.font = example_font

    # Headers, notes and the examples all stay visible while the organizer
    # scrolls their own rows.
    sheet.freeze_panes = f"A{2 + len(EXAMPLE_ROWS) + 1}"

    reference = workbook.create_sheet(SHEET_DISCIPLINES)
    reference.append(["Дисциплина", "Возраст", "Тип"])
    reference["A1"].font = Font(bold=True)
    reference["B1"].font = Font(bold=True)
    reference["C1"].font = Font(bold=True)
    for competition in competitions:
        reference.append(
            [
                competition.name,
                eligibility.describe_bounds(competition.min_age, competition.max_age) or "без ограничения",
                "команды" if competition.competition_type == "TEAM" else "лично",
            ]
        )
    reference.column_dimensions["A"].width = 32
    reference.column_dimensions["B"].width = 18
    reference.column_dimensions["C"].width = 12

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


# ------------------------------------------------------------------ the parse


@dataclass
class RawRow:
    """One data row, still text, before anything is checked."""

    row_number: int
    values: dict[str, str]


def parse_workbook(stream: BinaryIO) -> list[RawRow]:
    """Read the uploaded sheet into raw text rows.

    Columns are located by header text, so a sheet whose columns were reordered
    still imports. Only when no header is recognized does it fall back to
    :data:`IMPORT_COLUMNS` order, and a file matching neither is refused rather
    than silently read as gibberish.
    """
    try:
        workbook = load_workbook(stream, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 — openpyxl raises a zoo of format errors
        raise HTTPException(
            status_code=400,
            detail="Не удалось прочитать файл. Нужен .xlsx — старый .xls не поддерживается.",
        ) from None

    sheet = workbook[SHEET_ENTRIES] if SHEET_ENTRIES in workbook.sheetnames else workbook.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_by_norm = {_normalize(column.header_ru): column.key for column in IMPORT_COLUMNS}
    mapping: dict[int, str] = {}
    for index, cell in enumerate(rows[0]):
        key = header_by_norm.get(_normalize(_text(cell)))
        if key is not None:
            mapping[index] = key

    if mapping:
        body = rows[1:]
    else:
        # No recognizable header. Positional order is the only remaining
        # reading, and it is only worth trying if the first cell holds a name.
        if not _text(rows[0][0] if rows[0] else None):
            raise HTTPException(
                status_code=400,
                detail=(
                    "В файле не найдены заголовки колонок. Скачайте шаблон и заполните его — "
                    "первая строка должна называть колонки."
                ),
            )
        mapping = {index: column.key for index, column in enumerate(IMPORT_COLUMNS)}
        body = rows

    if len(body) > MAX_DATA_ROWS:
        raise HTTPException(
            status_code=400,
            detail=f"В файле больше {MAX_DATA_ROWS} строк — это не похоже на заявку одного турнира.",
        )

    parsed: list[RawRow] = []
    for offset, raw in enumerate(body):
        values = {column.key: "" for column in IMPORT_COLUMNS}
        for index, key in mapping.items():
            if index < len(raw):
                values[key] = _text(raw[index])
        # The note row of our own template, and any blank spacer row.
        if not any(values.values()):
            continue
        if values["full_name"] in {column.note for column in IMPORT_COLUMNS}:
            continue
        # A row the template filled in for illustration, left in place by an
        # organizer who did not delete it.
        if values["full_name"].startswith(EXAMPLE_MARKER):
            continue
        parsed.append(RawRow(row_number=offset + 2 if mapping else offset + 1, values=values))
    return parsed


# ------------------------------------------------------------- the validation


class ParticipantImportService:
    @staticmethod
    async def template(session: AsyncSession, tournament: Tournament) -> BytesIO:
        competitions = list(
            await session.scalars(
                select(Competition)
                .where(Competition.tournament_id == tournament.id)
                .order_by(Competition.created_at.asc())
            )
        )
        return build_template_workbook(competitions)

    @staticmethod
    async def validate(
        session: AsyncSession,
        tournament: Tournament,
        rows: Iterable[dict],
    ) -> dict:
        """Check every row against the database and report, writing nothing.

        Errors accumulate per row rather than short-circuiting: an organizer
        fixing a spreadsheet needs to see every problem in one pass, not
        discover the next one on each re-upload.
        """
        competitions = list(
            await session.scalars(
                select(Competition)
                .where(Competition.tournament_id == tournament.id)
                .order_by(Competition.created_at.asc())
            )
        )
        by_name = {_normalize(c.name): c for c in competitions}
        # A category name is accepted as an alias for the discipline that names
        # it, since the organizer's spreadsheet may well use either.
        categories = list(
            await session.scalars(
                select(TournamentCategory).where(TournamentCategory.tournament_id == tournament.id)
            )
        )
        category_to_competition = {
            _normalize(category.name): competition
            for category in categories
            for competition in competitions
            if competition.category_id == category.id
        }

        athletes = list(await session.scalars(select(Athlete)))
        athlete_by_nickname = {_normalize(a.nickname): a for a in athletes if a.nickname}

        existing = list(
            await session.scalars(
                select(Participant).where(Participant.tournament_id == tournament.id)
            )
        )
        existing_by_athlete = {
            (p.competition_id, p.athlete_id) for p in existing if p.athlete_id is not None
        }
        existing_by_name = {
            (p.competition_id, _normalize(p.display_name)) for p in existing if p.display_name
        }

        event_year = (
            tournament.start_date.year
            if tournament.start_date is not None
            else datetime.now(timezone.utc).year
        )

        seen_in_file: set[tuple[UUID | None, str]] = set()
        unknown_categories: set[str] = set()
        reported: list[dict] = []

        for index, raw in enumerate(rows, start=1):
            values = {column.key: _text(raw.get(column.key)) for column in IMPORT_COLUMNS}
            row_number = int(raw.get("row_number") or index)
            errors: list[dict] = []

            full_name = values["full_name"]
            if not full_name:
                errors.append(
                    {"code": "MISSING_NAME", "column": "full_name", "message": "Не указано ФИО"}
                )

            category_text = values["category"]
            competition = by_name.get(_normalize(category_text)) or category_to_competition.get(
                _normalize(category_text)
            )
            if competition is None:
                unknown_categories.add(category_text or "—")
                errors.append(
                    {
                        "code": "UNKNOWN_CATEGORY",
                        "column": "category",
                        "message": (
                            f"«{category_text}» не совпадает ни с одной дисциплиной турнира"
                            if category_text
                            else "Не указана дисциплина"
                        ),
                    }
                )

            birth_year: int | None = None
            if values["birth_year"]:
                try:
                    birth_year = int(values["birth_year"])
                except ValueError:
                    errors.append(
                        {
                            "code": "INVALID_BIRTH_YEAR",
                            "column": "birth_year",
                            "message": "Год рождения должен быть числом",
                        }
                    )
                else:
                    if not 1900 <= birth_year <= event_year:
                        errors.append(
                            {
                                "code": "INVALID_BIRTH_YEAR",
                                "column": "birth_year",
                                "message": f"Год рождения вне диапазона 1900–{event_year}",
                            }
                        )
                        birth_year = None

            seed: int | None = None
            if values["seed"]:
                try:
                    seed = int(values["seed"])
                except ValueError:
                    seed = None
                if seed is None or seed < 1:
                    errors.append(
                        {
                            "code": "INVALID_SEED",
                            "column": "seed",
                            "message": "Посев должен быть целым числом от 1",
                        }
                    )
                    seed = None

            reserve = False
            marked = _normalize(values["reserve"])
            if marked in RESERVE_YES:
                reserve = True
            elif marked not in RESERVE_NO:
                errors.append(
                    {
                        "code": "INVALID_RESERVE",
                        "column": "reserve",
                        "message": "В колонке «Запасной» пишут «да» или оставляют пусто",
                    }
                )

            # A fight name is the better match key: it is what a fighter is
            # actually known by, and what the roster shows.
            athlete = athlete_by_nickname.get(_normalize(values["fight_name"])) or (
                athlete_by_nickname.get(_normalize(full_name))
            )

            if competition is not None:
                if birth_year is None and values["birth_year"] == "" and athlete is not None:
                    birth_year = athlete.birth_year
                verdict = eligibility.check_age(
                    birth_year,
                    min_age=competition.min_age,
                    max_age=competition.max_age,
                    event_year=event_year,
                )
                if not verdict.ok:
                    errors.append(
                        {"code": verdict.code, "column": "birth_year", "message": verdict.message}
                    )

                key = (competition.id, _normalize(values["fight_name"] or full_name))
                if key in seen_in_file:
                    errors.append(
                        {
                            "code": "DUPLICATE_IN_FILE",
                            "column": "full_name",
                            "message": "Этот боец уже есть в файле в той же дисциплине",
                        }
                    )
                seen_in_file.add(key)

                already = (
                    athlete is not None and (competition.id, athlete.id) in existing_by_athlete
                ) or (competition.id, _normalize(full_name)) in existing_by_name
                if already:
                    errors.append(
                        {
                            "code": "DUPLICATE_IN_COMPETITION",
                            "column": "full_name",
                            "message": "Этот боец уже заявлен в этой дисциплине",
                        }
                    )

            reported.append(
                {
                    "row_number": row_number,
                    "full_name": full_name,
                    "fight_name": values["fight_name"] or None,
                    "city": values["city"] or None,
                    "club": values["club"] or None,
                    "category": category_text or None,
                    "birth_year": birth_year,
                    "seed": seed,
                    "reserve": reserve,
                    "competition_id": str(competition.id) if competition else None,
                    "competition_name": competition.name if competition else None,
                    "athlete_id": str(athlete.id) if athlete else None,
                    "athlete_display_name": athlete.nickname if athlete else None,
                    # Пусто — в сетке ФИО. Many fighters have no fight name.
                    "display_name": values["fight_name"] or full_name,
                    "errors": errors,
                    "valid": not errors,
                }
            )

        return {
            "tournament_id": str(tournament.id),
            "columns": [
                {
                    "key": column.key,
                    "header_ru": column.header_ru,
                    "required": column.required,
                    "note": column.note,
                }
                for column in IMPORT_COLUMNS
            ],
            "competitions": [
                {
                    "id": str(c.id),
                    "name": c.name,
                    "age_label": eligibility.describe_bounds(c.min_age, c.max_age),
                }
                for c in competitions
            ],
            "total_rows": len(reported),
            "valid_rows": sum(1 for row in reported if row["valid"]),
            "rows": reported,
            "unknown_categories": sorted(unknown_categories),
        }

    @staticmethod
    async def commit(
        session: AsyncSession,
        tournament: Tournament,
        rows: Iterable[dict],
        *,
        actor_id: UUID | None = None,
    ) -> dict:
        """Enter the submitted rows, refusing the batch if any of them is bad.

        Re-validates rather than trusting the preview: what arrives here is
        whatever the browser sent, possibly edited after the preview ran, and a
        client is never the authority on whether a row is admissible.

        All or nothing. A half-imported entry list is worse than a rejected one
        — the organizer cannot tell which half landed without reading the roster
        row by row.
        """
        from app.modules.tournaments.services.engine_service import TournamentEngineService

        report = await ParticipantImportService.validate(session, tournament, rows)
        if report["valid_rows"] != report["total_rows"]:
            raise HTTPException(status_code=400, detail=report)
        if report["total_rows"] == 0:
            raise HTTPException(status_code=400, detail="В заявке нет строк")

        per_competition: dict[str, int] = {}
        for row in report["rows"]:
            await TournamentEngineService.create_participant(
                session,
                competition_id=row["competition_id"],
                athlete_id=row["athlete_id"],
                display_name=row["display_name"],
                city=row["city"],
                club_name=row["club"],
                birth_year=row["birth_year"],
                seed=row["seed"],
                # A reserve is entered on the roster but stays out of the draw
                # until an organizer puts them in someone's place.
                status="RESERVE" if row.get("reserve") else "REGISTERED",
            )
            name = row["competition_name"] or row["competition_id"]
            per_competition[name] = per_competition.get(name, 0) + 1

        for competition_id in {row["competition_id"] for row in report["rows"]}:
            session.add(
                CompetitionEvent(
                    competition_id=UUID(competition_id),
                    event_type="PARTICIPANTS_IMPORTED",
                    description=f"Заявки загружены из файла: {report['total_rows']}",
                    payload={
                        "created": report["total_rows"],
                        "per_competition": per_competition,
                        "actor_id": str(actor_id) if actor_id else None,
                    },
                )
            )
        await session.flush()

        return {
            "tournament_id": str(tournament.id),
            "created": report["total_rows"],
            "per_competition": per_competition,
        }
